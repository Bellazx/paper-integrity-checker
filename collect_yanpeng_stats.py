#!/usr/bin/env python3
"""Collect metadata for Yan Peng papers and output as Markdown table."""
import json
import re
import time
import urllib.request
from pathlib import Path
from collections import defaultdict

import openpyxl

KB_DIR = Path("/opt/knowledge-base/data/paper-integrity/input/yanpeng")
CHECKER_DIR = Path("/opt/paper-integrity-checker/data/input/yanpeng")
EXCEL_PATH = CHECKER_DIR / "Yan_Peng_papers.xlsx"
OUTPUT_MD = CHECKER_DIR / "论文统计.md"


def _crossref_lookup(doi: str) -> dict:
    url = f"https://api.crossref.org/works/{doi}"
    req = urllib.request.Request(url)
    req.add_header("User-Agent", "PaperChecker/1.0 (mailto:paper-check@example.com)")
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
        item = data.get("message", {})
        journal = ""
        short = item.get("short-container-title", [])
        if short:
            journal = short[0]
        elif item.get("container-title"):
            journal = item["container-title"][0]
        year = ""
        pub_date = item.get("published-print") or item.get("published-online") or item.get("created")
        if pub_date:
            parts = pub_date.get("date-parts", [[]])[0]
            if parts:
                year = str(parts[0])
        return {"journal": journal, "year": year}
    except Exception:
        return {}


def load_excel():
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb.active
    data = {}
    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, 1).value or ""
        author = ws.cell(r, 2).value or ""
        journal = ws.cell(r, 3).value or ""
        doi = (ws.cell(r, 4).value or "").strip()
        year = ws.cell(r, 5).value or ""
        citations = ws.cell(r, 6).value or ""
        pdf_status = ws.cell(r, 7).value or ""
        if not doi:
            continue
        data[doi.lower()] = {
            "title": title.strip(),
            "author": author.strip(),
            "journal": journal.strip(),
            "year": str(year).strip() if year else "",
            "citations": str(citations).strip() if citations else "",
            "pdf_status": str(pdf_status).strip() if pdf_status else "",
        }
    return data


def load_json_papers():
    json_path = Path("/tmp/yanpeng_extract/yanpeng_papers/all_papers.json")
    if not json_path.exists():
        return {}
    with open(json_path) as f:
        papers = json.load(f)
    data = {}
    for p in papers:
        doi = p.get("doi", "").strip()
        if doi:
            venue = p.get("venue", "")
            if isinstance(venue, dict):
                venue = venue.get("info", {}).get("name", "")
            data[doi.lower()] = {
                "title": p.get("title", ""),
                "venue": venue,
                "year": str(p.get("year", "")),
                "authors": p.get("authors", []),
            }
    return data


def scan_papers():
    excel_data = load_excel()
    json_data = load_json_papers()
    print(f"Excel entries: {len(excel_data)}, JSON entries: {len(json_data)}")

    all_dois = set()
    for d in excel_data:
        all_dois.add(d)
    for d in json_data:
        all_dois.add(d)

    crossref_count = 0
    papers = []

    for doi_lower in sorted(all_dois):
        excel_info = excel_data.get(doi_lower, {})
        json_info = json_data.get(doi_lower, {})

        doi = doi_lower
        for d in excel_data:
            if d.lower() == doi_lower:
                doi = (list(excel_data.keys())[[k.lower() for k in excel_data.keys()].index(doi_lower)])
                break

        title = excel_info.get("title") or json_info.get("title", "")
        journal = excel_info.get("journal") or json_info.get("venue", "")
        year = excel_info.get("year") or json_info.get("year", "")
        author = excel_info.get("author", "")

        dirname = doi.replace("/", "_")
        checker_dir = CHECKER_DIR / dirname
        kb_dir = KB_DIR / dirname

        has_pdf = False
        if checker_dir.exists():
            has_pdf = any(checker_dir.glob("*.pdf"))
        if not has_pdf and kb_dir.exists():
            has_pdf = any(kb_dir.glob("*.pdf"))

        has_supp = False
        supp_types = []
        for base in [checker_dir, kb_dir]:
            if not base.exists():
                continue
            for sub in ("supplementary", "source_data", "source data"):
                sd = base / sub
                if sd.exists() and any(sd.iterdir()):
                    has_supp = True
                    for f in sd.iterdir():
                        if f.suffix.lower() in (".xlsx", ".xls", ".csv", ".zip", ".rar"):
                            supp_types.append(f.suffix.lower())

        if (not journal or not year) and doi.startswith("10.") and crossref_count < 100:
            cr = _crossref_lookup(doi)
            crossref_count += 1
            if crossref_count % 10 == 0:
                print(f"  CrossRef lookups: {crossref_count}")
            time.sleep(0.15)
            if cr.get("journal") and not journal:
                journal = cr["journal"]
            if cr.get("year") and not year:
                year = cr["year"]

        supp_str = "有" if has_supp else "无"
        if supp_types:
            unique_types = sorted(set(supp_types))
            supp_str = "有 (" + ", ".join(unique_types) + ")"

        papers.append({
            "doi": doi,
            "title": title,
            "author": author,
            "journal": journal,
            "year": year,
            "has_pdf": has_pdf,
            "has_supp": has_supp,
            "supp_str": supp_str,
        })

    return papers


def write_md(papers):
    with_pdf = sum(1 for p in papers if p["has_pdf"])
    without_pdf = len(papers) - with_pdf
    with_supp = sum(1 for p in papers if p["has_supp"])

    journals = defaultdict(int)
    years = defaultdict(int)
    for p in papers:
        if p["journal"]:
            journals[p["journal"]] += 1
        if p["year"]:
            years[p["year"]] += 1

    lines = []
    lines.append("# 彭岩论文统计\n")
    lines.append(f"统计日期：2026-05-25\n")
    lines.append(f"## 概览\n")
    lines.append(f"- 论文总数：**{len(papers)}** 篇")
    lines.append(f"- 有PDF原文：**{with_pdf}** 篇")
    lines.append(f"- 无PDF原文：**{without_pdf}** 篇")
    lines.append(f"- 有补充材料：**{with_supp}** 篇")
    lines.append(f"- 无补充材料：**{len(papers) - with_supp}** 篇\n")

    lines.append("### 期刊分布\n")
    lines.append("| 期刊 | 篇数 |")
    lines.append("|------|------|")
    for j, c in sorted(journals.items(), key=lambda x: -x[1]):
        lines.append(f"| {j} | {c} |")
    lines.append("")

    lines.append("### 年份分布\n")
    lines.append("| 年份 | 篇数 |")
    lines.append("|------|------|")
    for y in sorted(years.keys()):
        lines.append(f"| {y} | {years[y]} |")
    lines.append("")

    lines.append("## 论文详细列表\n")
    lines.append("| 序号 | DOI | 期刊 | 年份 | PDF | 补充材料 |")
    lines.append("|------|-----|------|------|-----|---------|")
    for i, p in enumerate(papers, 1):
        doi = p["doi"]
        journal = p["journal"] or "-"
        year = p["year"] or "-"
        pdf = "有" if p["has_pdf"] else "无"
        supp = p["supp_str"]
        lines.append(f"| {i} | {doi} | {journal} | {year} | {pdf} | {supp} |")

    lines.append("")
    content = "\n".join(lines)
    OUTPUT_MD.write_text(content, encoding="utf-8")
    print(f"\nOutput written to {OUTPUT_MD}")
    print(f"Total papers: {len(papers)}")
    print(f"With PDF: {with_pdf}, Without PDF: {without_pdf}")
    print(f"With supplementary: {with_supp}")


if __name__ == "__main__":
    papers = scan_papers()
    write_md(papers)
