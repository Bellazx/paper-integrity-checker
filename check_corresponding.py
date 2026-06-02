#!/usr/bin/env python3
"""Check how many papers have Zhang Wanbin as corresponding author."""
import json
import re
import time
import urllib.request
from pathlib import Path

import openpyxl

KB_DIR = Path("/opt/knowledge-base/data/paper-integrity/input/zhangwanbin")
EXCEL_PATH = Path("/opt/paper-integrity-checker/data/input/zhangwanbin/学术数据-Excel-20260525 (1).xlsx")

ZHANG_PATTERNS = [
    "zhang, wanbin", "zhang,wanbin", "zhang wanbin", "wanbin zhang",
    "万斌", "张万斌",
]


def _is_zhang_wanbin(name: str) -> bool:
    name_lower = name.lower().strip()
    for p in ZHANG_PATTERNS:
        if p in name_lower:
            return True
    name_clean = re.sub(r'\(.*?\)', '', name_lower).strip().rstrip(',')
    for p in ZHANG_PATTERNS:
        if p in name_clean:
            return True
    return False


def _crossref_authors(doi: str) -> list[str] | None:
    url = f"https://api.crossref.org/works/{doi}"
    req = urllib.request.Request(url)
    req.add_header("User-Agent", "PaperChecker/1.0 (mailto:paper-check@example.com)")
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
        item = data.get("message", {})
        authors = item.get("author", [])
        names = []
        for a in authors:
            family = a.get("family", "")
            given = a.get("given", "")
            if family:
                names.append(f"{family}, {given}".strip().rstrip(","))
            elif a.get("name"):
                names.append(a["name"])
        return names if names else None
    except Exception:
        return None


def load_excel_authors() -> dict:
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb.active
    data = {}
    for r in range(2, ws.max_row + 1):
        author = ws.cell(r, 2).value or ""
        doi = (ws.cell(r, 6).value or "").strip()
        if not doi or not author:
            continue
        authors = [re.sub(r'\(.*?\)', '', a).strip().rstrip(',') for a in author.split(';') if a.strip()]
        data[doi] = authors
    return data


def main():
    excel_authors = load_excel_authors()

    dirs = sorted([d for d in KB_DIR.iterdir() if d.is_dir()])
    seen = set()
    results = {"corresponding": [], "not_corresponding": [], "unknown": []}
    cr_count = 0

    for paper_dir in dirs:
        if not any(paper_dir.glob("*.pdf")):
            continue

        doi_file = paper_dir / "doi.txt"
        if doi_file.exists():
            doi = doi_file.read_text(encoding="utf-8-sig").strip()
        else:
            doi = paper_dir.name.replace("_", "/", 1)

        if doi.lower() in seen:
            continue
        seen.add(doi.lower())

        if not doi.startswith("10."):
            results["unknown"].append({"doi": doi, "reason": "no_doi"})
            continue

        authors = excel_authors.get(doi)
        source = "excel"

        if not authors:
            authors = _crossref_authors(doi)
            source = "crossref"
            cr_count += 1
            if cr_count % 20 == 0:
                print(f"  CrossRef lookups: {cr_count}")
            time.sleep(0.15)

        if not authors:
            results["unknown"].append({"doi": doi, "reason": "no_author_data"})
            continue

        zhang_positions = [i for i, a in enumerate(authors) if _is_zhang_wanbin(a)]

        if not zhang_positions:
            results["unknown"].append({"doi": doi, "reason": "zhang_not_found", "authors": authors[-3:], "source": source})
            continue

        is_last = (len(authors) - 1) in zhang_positions
        # Also check second-to-last for co-corresponding
        is_second_last = (len(authors) - 2) in zhang_positions if len(authors) >= 2 else False

        if is_last:
            results["corresponding"].append({"doi": doi, "position": "last", "total": len(authors)})
        elif is_second_last and len(authors) >= 3:
            results["corresponding"].append({"doi": doi, "position": "second_last", "total": len(authors)})
        else:
            pos = zhang_positions[0]
            results["not_corresponding"].append({
                "doi": doi, "position": pos, "total": len(authors),
                "last_author": authors[-1] if authors else "",
            })

    print(f"\n{'='*60}")
    print(f"张万斌通讯作者统计")
    print(f"{'='*60}")
    print(f"论文总数: {len(seen)}")
    print(f"通讯作者（末位/倒数第二位）: {len(results['corresponding'])}")
    print(f"非通讯作者: {len(results['not_corresponding'])}")
    print(f"无法判断: {len(results['unknown'])}")

    print(f"\n--- 非通讯作者 ({len(results['not_corresponding'])}) ---")
    for r in results["not_corresponding"]:
        print(f"  {r['doi']}: 张万斌在第{r['position']+1}位/{r['total']}人, 末位={r['last_author']}")

    print(f"\n--- 无法判断 ({len(results['unknown'])}) ---")
    for r in results["unknown"]:
        reason = r["reason"]
        if reason == "no_doi":
            print(f"  {r['doi'][:50]}: 无DOI")
        elif reason == "no_author_data":
            print(f"  {r['doi']}: 无作者数据")
        elif reason == "zhang_not_found":
            print(f"  {r['doi']}: 作者列表中未找到张万斌 (末尾: {r['authors'][-1] if r.get('authors') else '?'}, src={r.get('source','')})")


if __name__ == "__main__":
    main()
