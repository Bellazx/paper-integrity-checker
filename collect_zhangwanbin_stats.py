#!/usr/bin/env python3
"""Collect metadata for Zhang Wanbin papers and output as Markdown table."""
import os
import re
import json
import time
import urllib.request
from pathlib import Path
from collections import defaultdict

import openpyxl

KB_DIR = Path("/opt/knowledge-base/data/paper-integrity/input/zhangwanbin")
CHECKER_DIR = Path("/opt/paper-integrity-checker/data/input/zhangwanbin")
EXCEL_PATH = CHECKER_DIR / "学术数据-Excel-20260525 (1).xlsx"
OUTPUT_MD = CHECKER_DIR / "论文统计.md"

DOI_JOURNAL_MAP = {
    "10.1002/adsc": "Adv. Synth. Catal.",
    "10.1002/anie": "Angew. Chem. Int. Ed.",
    "10.1002/ange": "Angew. Chem.",
    "10.1002/chem": "Chem. Eur. J.",
    "10.1002/cctc": "ChemCatChem",
    "10.1002/advs": "Adv. Sci.",
    "10.1002/ajoc": "Asian J. Org. Chem.",
    "10.1002/ejoc": "Eur. J. Org. Chem.",
    "10.1002/ejic": "Eur. J. Inorg. Chem.",
    "10.1002/cssc": "ChemSusChem",
    "10.1002/cbic": "ChemBioChem",
    "10.1002/cmdc": "ChemMedChem",
    "10.1021/jacs": "J. Am. Chem. Soc.",
    "10.1021/ja": "J. Am. Chem. Soc.",
    "10.1021/acs.orglett": "Org. Lett.",
    "10.1021/ol": "Org. Lett.",
    "10.1021/acscatal": "ACS Catal.",
    "10.1021/acs.accounts": "Acc. Chem. Res.",
    "10.1021/acs.joc": "J. Org. Chem.",
    "10.1021/jo": "J. Org. Chem.",
    "10.1021/acs.chemrev": "Chem. Rev.",
    "10.1021/cr": "Chem. Rev.",
    "10.1021/acs.oprd": "Org. Process Res. Dev.",
    "10.1021/acs.jpcb": "J. Phys. Chem. B",
    "10.1021/acs.organomet": "Organometallics",
    "10.1021/jm": "J. Med. Chem.",
    "10.1021/la": "Langmuir",
    "10.1021/prechem": "JACS Au",
    "10.1021/cs": "ACS Catal.",
    "10.1038/s41467": "Nat. Commun.",
    "10.1038/ncomms": "Nat. Commun.",
    "10.1038/s41929": "Nat. Catal.",
    "10.1038/s41557": "Nat. Chem.",
    "10.1038/nchem": "Nat. Chem.",
    "10.1038/s41586": "Nature",
    "10.1038/s42004": "Commun. Chem.",
    "10.1039/c": "RSC (various)",
    "10.1039/d": "RSC (various)",
    "10.1039/b": "RSC (various)",
    "10.1016/j.tetlet": "Tetrahedron Lett.",
    "10.1016/j.tet": "Tetrahedron",
    "10.1016/j.tetasy": "Tetrahedron: Asymmetry",
    "10.1016/j.cclet": "Chin. Chem. Lett.",
    "10.1016/j.jorganchem": "J. Organomet. Chem.",
    "10.1016/j.chempr": "Chem",
    "10.1016/j.checat": "Chem Catal.",
    "10.1016/j.scib": "Sci. Bull.",
    "10.1016/j.jcat": "J. Catal.",
    "10.1016/j.molcata": "J. Mol. Catal. A",
    "10.1016/j.apsusc": "Appl. Surf. Sci.",
    "10.1016/j.bmcl": "Bioorg. Med. Chem. Lett.",
    "10.1016/j.dyepig": "Dyes Pigm.",
    "10.1016/j.phrs": "Pharmacol. Res.",
    "10.1016/j.polymer": "Polymer",
    "10.1016/j.polymertesting": "Polym. Test.",
    "10.1016/j.fuel": "Fuel",
    "10.1016/S0040-4039": "Tetrahedron Lett.",
    "10.1016/S0040-4020": "Tetrahedron",
    "10.1016/S0957-4166": "Tetrahedron: Asymmetry",
    "10.1016/s0040-4039": "Tetrahedron Lett.",
    "10.1016/s0957-4166": "Tetrahedron: Asymmetry",
    "10.1016/0040-4039": "Tetrahedron Lett.",
    "10.1016/0957-4166": "Tetrahedron: Asymmetry",
    "10.1016j/fuel": "Fuel",
    "10.1126/science": "Science",
    "10.1055/s": "Synlett/Synthesis",
    "10.1055/a": "Synlett/Synthesis",
    "10.6023/cjoc": "有机化学",
    "10.6023/A": "有机化学",
    "10.6023/a": "有机化学",
    "10.3866/PKU": "大学化学",
    "10.5650/jos": "J. Oleo Sci.",
    "10.5517/cc": "CSD Communication",
    "10.1246/bcsj": "Bull. Chem. Soc. Jpn.",
    "10.1246/cl": "Chem. Lett.",
    "10.1007/s11426": "Sci. China Chem.",
    "10.1007/s10562": "Catal. Lett.",
    "10.1007/s00253": "Appl. Microbiol. Biotechnol.",
    "10.1007/s10311": "Environ. Chem. Lett.",
    "10.1007/s10870": "J. Chem. Crystallogr.",
    "10.31635/ccschem": "CCS Chem.",
    "10.15227/orgsyn": "Org. Synth.",
    "10.1360/N032": "中国科学",
    "10.1360/032": "中国科学",
    "10.3969/j": "中国期刊",
    "10.16009/j": "中国期刊",
    "10.16522/j": "中国期刊",
    "10.1111/jfpp": "J. Food Process. Preserv.",
    "10.1371/journal": "PLoS ONE",
    "10.3144/express": "J. Oleo Sci.",
}

RSC_JOURNAL_MAP = {
    "cc": "Chem. Commun.",
    "gc": "Green Chem.",
    "sc": "Chem. Sci.",
    "ob": "Org. Biomol. Chem.",
    "cy": "Catal. Sci. Technol.",
    "dt": "Dalton Trans.",
    "nj": "New J. Chem.",
    "ra": "RSC Adv.",
    "qo": "Org. Chem. Front.",
    "re": "Chem. Soc. Rev.",
    "c3cc": "Chem. Commun.",
    "c4cc": "Chem. Commun.",
    "c5cc": "Chem. Commun.",
    "c6cc": "Chem. Commun.",
    "c7cc": "Chem. Commun.",
    "c8cc": "Chem. Commun.",
    "c9cc": "Chem. Commun.",
    "c3sc": "Chem. Sci.",
    "c4sc": "Chem. Sci.",
    "c5sc": "Chem. Sci.",
    "c6sc": "Chem. Sci.",
    "c7sc": "Chem. Sci.",
    "c8sc": "Chem. Sci.",
    "c9sc": "Chem. Sci.",
    "c0sc": "Chem. Sci.",
    "c3ob": "Org. Biomol. Chem.",
    "c4ob": "Org. Biomol. Chem.",
    "c5ob": "Org. Biomol. Chem.",
    "c6qo": "Org. Chem. Front.",
    "c7qo": "Org. Chem. Front.",
    "c8qo": "Org. Chem. Front.",
    "c9qo": "Org. Chem. Front.",
    "c5cy": "Catal. Sci. Technol.",
    "c6cy": "Catal. Sci. Technol.",
    "c7cy": "Catal. Sci. Technol.",
    "d0sc": "Chem. Sci.",
    "d0cc": "Chem. Commun.",
    "d0qo": "Org. Chem. Front.",
    "d0cy": "Catal. Sci. Technol.",
    "d1sc": "Chem. Sci.",
    "d1cc": "Chem. Commun.",
    "d1qo": "Org. Chem. Front.",
    "d1cy": "Catal. Sci. Technol.",
    "d2sc": "Chem. Sci.",
    "d2cc": "Chem. Commun.",
    "d2qo": "Org. Chem. Front.",
    "d2cy": "Catal. Sci. Technol.",
    "d3sc": "Chem. Sci.",
    "d3cc": "Chem. Commun.",
    "d3qo": "Org. Chem. Front.",
    "d3cy": "Catal. Sci. Technol.",
    "d4sc": "Chem. Sci.",
    "d4cc": "Chem. Commun.",
    "d4qo": "Org. Chem. Front.",
    "d4cy": "Catal. Sci. Technol.",
    "d5sc": "Chem. Sci.",
    "d5cc": "Chem. Commun.",
    "d5qo": "Org. Chem. Front.",
    "d5cy": "Catal. Sci. Technol.",
}


def _doi_to_journal(doi: str) -> str:
    if not doi:
        return ""
    if doi.startswith("10.1039/"):
        suffix = doi.split("/", 1)[1].lower()
        for prefix, journal in sorted(RSC_JOURNAL_MAP.items(), key=lambda x: -len(x[0])):
            if suffix.startswith(prefix):
                return journal
        if "cs" in suffix:
            return "Chem. Soc. Rev."
        if "ob" in suffix:
            return "Org. Biomol. Chem."
        if "cc" in suffix:
            return "Chem. Commun."
        if "gc" in suffix:
            return "Green Chem."

    for prefix, journal in sorted(DOI_JOURNAL_MAP.items(), key=lambda x: -len(x[0])):
        if doi.startswith(prefix):
            return journal
    return ""


def _doi_to_year(doi: str) -> str:
    if not doi:
        return ""
    m = re.search(r'[._-]((?:19|20)\d{2})', doi)
    if m:
        return m.group(1)
    m = re.search(r'(?:20)(\d{2})\d{3,}', doi.split("/")[-1])
    if m:
        y = int("20" + m.group(1))
        if 2000 <= y <= 2026:
            return str(y)
    return ""


def _crossref_lookup(doi: str) -> dict:
    url = f"https://api.crossref.org/works/{doi}"
    req = urllib.request.Request(url)
    req.add_header("User-Agent", "PaperChecker/1.0 (mailto:paper-check@example.com)")
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
        item = data.get("message", {})
        journal = ""
        titles = item.get("container-title", [])
        if titles:
            journal = titles[0]
        short = item.get("short-container-title", [])
        if short:
            journal = short[0]

        year = ""
        pub_date = item.get("published-print") or item.get("published-online") or item.get("created")
        if pub_date:
            parts = pub_date.get("date-parts", [[]])[0]
            if parts:
                year = str(parts[0])

        return {"journal": journal, "year": year}
    except Exception:
        return {}


def load_excel():
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb.active
    data = {}
    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, 1).value or ""
        author = ws.cell(r, 2).value or ""
        impact_factor = ws.cell(r, 3).value or ""
        unit = ws.cell(r, 4).value or ""
        volume = ws.cell(r, 5).value or ""
        doi = ws.cell(r, 6).value or ""
        issn = ws.cell(r, 7).value or ""
        pub_date = ws.cell(r, 8).value or ""
        if not doi:
            continue
        doi = doi.strip()
        year = ""
        if pub_date:
            pub_str = str(pub_date)
            m = re.search(r'((?:19|20)\d{2})', pub_str)
            if m:
                year = m.group(1)
        data[doi] = {
            "title": title.strip() if title else "",
            "author": author.strip() if author else "",
            "impact_factor": str(impact_factor).strip() if impact_factor else "",
            "unit": unit.strip() if unit else "",
            "volume": volume.strip() if volume else "",
            "issn": issn.strip() if issn else "",
            "year": year,
        }
    return data


def scan_papers():
    papers = []
    dirs = sorted([d for d in KB_DIR.iterdir() if d.is_dir()])

    excel_data = load_excel()
    print(f"Excel entries: {len(excel_data)}")

    crossref_cache = {}
    crossref_count = 0
    seen_dois = set()

    for paper_dir in dirs:
        doi_file = paper_dir / "doi.txt"
        if doi_file.exists():
            doi = doi_file.read_text(encoding="utf-8-sig").strip()
        else:
            doi = paper_dir.name.replace("_", "/", 1)

        has_pdf = any(paper_dir.glob("*.pdf"))
        if not has_pdf:
            continue

        if doi.lower() in seen_dois:
            continue
        seen_dois.add(doi.lower())

        checker_dir = CHECKER_DIR / paper_dir.name
        has_supp = False
        supp_types = []
        if checker_dir.exists():
            for sub in ("supplementary", "source data", "source_data"):
                sd = checker_dir / sub
                if sd.exists() and any(sd.iterdir()):
                    has_supp = True
                    for f in sd.iterdir():
                        if f.suffix.lower() in (".xlsx", ".xls", ".csv", ".zip", ".rar", ".docx", ".doc", ".pdf"):
                            supp_types.append(f.suffix.lower())
            if not has_supp:
                data_files = list(checker_dir.rglob("*.xlsx")) + list(checker_dir.rglob("*.xls")) + list(checker_dir.rglob("*.csv"))
                data_files = [f for f in data_files if f.name != paper_dir.name + ".xlsx"]
                if data_files:
                    has_supp = True
                    supp_types = [f.suffix.lower() for f in data_files]

        excel_info = excel_data.get(doi, {})
        impact_factor = excel_info.get("impact_factor", "")
        year = excel_info.get("year", "")
        journal = _doi_to_journal(doi)

        if not year:
            year = _doi_to_year(doi)

        if (not journal or not year) and doi.startswith("10.") and crossref_count < 300:
            if doi not in crossref_cache:
                cr = _crossref_lookup(doi)
                crossref_cache[doi] = cr
                crossref_count += 1
                if crossref_count % 20 == 0:
                    print(f"  CrossRef lookups: {crossref_count}")
                time.sleep(0.15)
            cr = crossref_cache.get(doi, {})
            if cr.get("journal") and not journal:
                journal = cr["journal"]
            if cr.get("year") and not year:
                year = cr["year"]

        supp_str = "有" if has_supp else "无"
        if supp_types:
            unique_types = sorted(set(supp_types))
            supp_str = "有 (" + ", ".join(unique_types) + ")"

        papers.append({
            "doi": doi,
            "journal": journal,
            "impact_factor": impact_factor,
            "year": year,
            "has_supp": has_supp,
            "supp_str": supp_str,
            "title": excel_info.get("title", ""),
        })

    return papers


def write_md(papers):
    with_supp = sum(1 for p in papers if p["has_supp"])
    without_supp = len(papers) - with_supp

    journals = defaultdict(int)
    years = defaultdict(int)
    for p in papers:
        if p["journal"]:
            journals[p["journal"]] += 1
        if p["year"]:
            years[p["year"]] += 1

    lines = []
    lines.append("# 张万斌论文统计\n")
    lines.append(f"统计日期：2026-05-25\n")
    lines.append(f"## 概览\n")
    lines.append(f"- 论文总数：**{len(papers)}** 篇")
    lines.append(f"- 有补充材料：**{with_supp}** 篇")
    lines.append(f"- 无补充材料：**{without_supp}** 篇\n")

    lines.append("### 期刊分布\n")
    lines.append("| 期刊 | 篇数 |")
    lines.append("|------|------|")
    for j, c in sorted(journals.items(), key=lambda x: -x[1]):
        lines.append(f"| {j} | {c} |")
    if any(not p["journal"] for p in papers):
        unknown = sum(1 for p in papers if not p["journal"])
        lines.append(f"| _(未识别)_ | {unknown} |")
    lines.append("")

    lines.append("### 年份分布\n")
    lines.append("| 年份 | 篇数 |")
    lines.append("|------|------|")
    for y in sorted(years.keys()):
        lines.append(f"| {y} | {years[y]} |")
    if any(not p["year"] for p in papers):
        unknown_y = sum(1 for p in papers if not p["year"])
        lines.append(f"| _(未知)_ | {unknown_y} |")
    lines.append("")

    lines.append("## 论文详细列表\n")
    lines.append("| 序号 | DOI | 期刊 | 影响因子 | 年份 | 补充材料 |")
    lines.append("|------|-----|------|---------|------|---------|")
    for i, p in enumerate(papers, 1):
        doi = p["doi"]
        journal = p["journal"] or "-"
        impact = p["impact_factor"] or "-"
        year = p["year"] or "-"
        supp = p["supp_str"]
        lines.append(f"| {i} | {doi} | {journal} | {impact} | {year} | {supp} |")

    lines.append("")
    content = "\n".join(lines)
    OUTPUT_MD.write_text(content, encoding="utf-8")
    print(f"Output written to {OUTPUT_MD}")
    print(f"Total papers: {len(papers)}")
    print(f"With supplementary: {with_supp}")
    print(f"Without supplementary: {without_supp}")


if __name__ == "__main__":
    papers = scan_papers()
    write_md(papers)
