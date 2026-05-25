"""Excel batch metadata loading and matching utilities.

Each batch folder (e.g., data/input/20260520-649/) may contain an Excel file from
InCites with authoritative metadata (title, journal, authors). This module loads
that data and merges it into pipeline findings before DB insertion.
"""
import logging
from pathlib import Path

log = logging.getLogger(__name__)


def find_batch_excel(batch_dir: Path) -> Path | None:
    """Find the Excel metadata file in a batch input directory."""
    batch_dir = Path(batch_dir)
    if not batch_dir.is_dir():
        return None
    xlsx_files = list(batch_dir.glob("*.xlsx"))
    if not xlsx_files:
        return None
    if len(xlsx_files) > 1:
        log.info("Multiple xlsx files in %s, using first: %s", batch_dir, xlsx_files[0].name)
    return xlsx_files[0]


def load_batch_excel(excel_path: Path) -> dict[str, dict]:
    """Load DOI -> metadata mapping from a batch Excel file.

    Returns dict keyed by DOI with values:
        {"title": str, "authors": str, "source": str}
    """
    import openpyxl

    excel_path = Path(excel_path)
    log.info("Loading batch Excel: %s", excel_path.name)

    wb = openpyxl.load_workbook(str(excel_path), read_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    doi_idx = _find_column(headers, ["DOI"])
    title_idx = _find_column(headers, ["Article Title", "Title"])
    source_idx = _find_column(headers, ["Source", "Journal", "source"])
    authors_idx = _find_column(headers, ["Authors", "Author"])

    if doi_idx is None:
        log.warning("No DOI column found in Excel, skipping")
        wb.close()
        return {}

    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        doi = str(row[doi_idx]).strip() if row[doi_idx] else ""
        if not doi:
            continue
        data[doi] = {
            "title": str(row[title_idx]).strip() if title_idx is not None and row[title_idx] else "",
            "source": str(row[source_idx]).strip() if source_idx is not None and row[source_idx] else "",
            "authors": str(row[authors_idx]).strip() if authors_idx is not None and row[authors_idx] else "",
        }

    wb.close()
    log.info("Loaded %d DOI entries from Excel", len(data))
    return data


def merge_excel_metadata(findings: dict, excel_data: dict) -> dict:
    """Compare and override findings metadata with Excel data before DB insertion.

    Overwrites title and journal from Excel (more reliable than PDF extraction).
    Does NOT override authors_full/affiliations (LLM extraction is richer).
    """
    if not excel_data:
        return findings

    doi = findings.get("paper", {}).get("doi", "")
    if not doi or doi not in excel_data:
        return findings

    excel = excel_data[doi]
    paper = findings["paper"]

    if excel["title"]:
        old_title = paper.get("title", "")
        paper["title"] = excel["title"]
        if old_title != excel["title"]:
            log.debug("Title overridden by Excel for %s", doi)

    if excel["source"]:
        old_journal = paper.get("journal", "")
        paper["journal"] = excel["source"]
        if old_journal != excel["source"]:
            log.debug("Journal overridden by Excel for %s", doi)

    return findings


def _find_column(headers: list, candidates: list[str]) -> int | None:
    """Find column index matching any candidate name (case-insensitive)."""
    for i, h in enumerate(headers):
        if h and any(c.lower() == str(h).lower() for c in candidates):
            return i
    return None
